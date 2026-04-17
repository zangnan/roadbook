"""路书 Excel 导出模块"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_roadbook_to_excel(roadbook_data, output_path):
    """
    将路书数据导出为 Excel 文件

    Args:
        roadbook_data: 路书 JSON 数据
        output_path: 输出文件路径
    """
    wb = Workbook()

    # 移除默认 sheet
    wb.remove(wb.active)

    # 创建各 Sheet
    _create_overview_sheet(wb, roadbook_data)
    _create_daily_detail_sheet(wb, roadbook_data)
    _create_scenic_guides_sheet(wb, roadbook_data)
    _create_budget_sheet(wb, roadbook_data)
    _create_checklist_sheet(wb, roadbook_data)

    # 方案 A: 新增独立的「路线站点」Sheet
    _create_route_stops_sheet(wb, roadbook_data)

    # 方案 B: 在每日详情中合并路线站点（新增一列）
    _update_daily_detail_with_stops(wb, roadbook_data)

    # 保存文件
    wb.save(output_path)


def _create_overview_sheet(wb, roadbook_data):
    """行程概览 Sheet"""
    ws = wb.create_sheet("行程概览")

    # 标题样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["日期", "星期", "出发地", "目的地", "里程(km)", "时长(h)", "海拔(m)", "路线"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    daily_itinerary = roadbook_data.get('daily_itinerary', [])
    for row_idx, day in enumerate(daily_itinerary, 2):
        date_str = day.get('date', '')
        # 计算星期
        weekday = ""
        if date_str:
            try:
                weekday = ["一", "二", "三", "四", "五", "六", "日"][
                    datetime.strptime(date_str, '%Y-%m-%d').weekday()
                ]
            except:
                pass

        row_data = [
            date_str,
            weekday,
            day.get('origin', ''),
            day.get('destination', ''),
            day.get('distance_km', ''),
            day.get('duration_hours', ''),
            day.get('elevation_m', ''),
            day.get('route', '')
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    col_widths = [12, 6, 10, 10, 10, 8, 10, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _create_daily_detail_sheet(wb, roadbook_data):
    """每日详情 Sheet"""
    ws = wb.create_sheet("每日详情")

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["天数", "日期", "出发地", "目的地", "里程(km)", "时长(h)", "亮点", "住宿", "美食", "tips"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    daily_itinerary = roadbook_data.get('daily_itinerary', [])
    for row_idx, day in enumerate(daily_itinerary, 2):
        highlights = "、".join(day.get('highlights', [])[:3]) if day.get('highlights') else ""
        food = "、".join(day.get('food', [])[:3]) if day.get('food') else ""
        tips = "；".join(day.get('tips', [])[:2]) if day.get('tips') else ""

        row_data = [
            day.get('day_number', ''),
            day.get('date', ''),
            day.get('origin', ''),
            day.get('destination', ''),
            day.get('distance_km', ''),
            day.get('duration_hours', ''),
            highlights,
            day.get('accommodation', ''),
            food,
            tips
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 设置列宽
    col_widths = [6, 12, 10, 10, 10, 8, 30, 20, 25, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 设置行高
    for row in range(2, len(daily_itinerary) + 2):
        ws.row_dimensions[row].height = 45


def _update_daily_detail_with_stops(wb, roadbook_data):
    """方案 B: 在每日详情中合并路线站点"""
    ws = wb["每日详情"]

    # 站点类型图标映射
    type_icons = {
        'start': '🚗',
        'gas': '⛽',
        'transit': '🏙',
        'scenic': '🏞',
        'end': '🏁',
        'food': '🍜',
        'accommodation': '🏨'
    }

    # 从 daily_itinerary 中获取每个 day 的 route_stops
    daily_itinerary = roadbook_data.get('daily_itinerary', [])

    # 在每日详情最后新增「途经站点」列
    stops_col = 11

    # 添加表头
    cell = ws.cell(row=1, column=stops_col, value='途经站点')
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 填充每天的站点
    for row_idx, day in enumerate(daily_itinerary, 2):
        # route_stops 嵌套在 daily_itinerary 每个 day 里面
        stops = day.get('route_stops', [])

        stops_text = ""
        for stop in stops:
            icon = type_icons.get(stop.get('type'), '📍')
            name = stop.get('name', '')
            stops_text += f"{icon}{name} "

        cell = ws.cell(row=row_idx, column=stops_col, value=stops_text.strip())
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 设置列宽
    ws.column_dimensions[get_column_letter(stops_col)].width = 30


def _create_route_stops_sheet(wb, roadbook_data):
    """方案 A: 路线站点 Sheet"""
    ws = wb.create_sheet("路线站点")

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["天数", "类型", "站点名称", "位置", "说明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 站点类型映射
    type_names = {
        'start': '出发',
        'gas': '加油/服务区',
        'transit': '途经城市',
        'scenic': '景点',
        'end': '终点',
        'food': '美食',
        'accommodation': '住宿'
    }

    # 从 daily_itinerary 中收集所有 route_stops
    daily_itinerary = roadbook_data.get('daily_itinerary', [])

    row_idx = 2
    for day in daily_itinerary:
        day_number = day.get('day_number', '')
        stops = day.get('route_stops', [])

        for stop in stops:
            stop_type = stop.get('type', 'transit')
            type_name = type_names.get(stop_type, stop_type)

            row_data = [
                day_number,
                type_name,
                stop.get('name', ''),
                stop.get('location', ''),
                stop.get('description', '')
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            row_idx += 1

    # 设置列宽
    col_widths = [6, 10, 18, 22, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 设置行高
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 30


def _create_scenic_guides_sheet(wb, roadbook_data):
    """景点攻略 Sheet"""
    ws = wb.create_sheet("景点攻略")

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["景点名称", "位置", "开放时间", "建议游览时长", "门票信息", "详细介绍", "游览贴士"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    scenic_guides = roadbook_data.get('scenic_guides', {})
    row_idx = 2
    for spot_name, guide in scenic_guides.items():
        tips = "；".join(guide.get('tips', [])[:3]) if guide.get('tips') else ""

        row_data = [
            guide.get('name', spot_name),
            guide.get('location', ''),
            guide.get('opening_hours', ''),
            guide.get('recommended_visit_time', ''),
            guide.get('ticket_info', ''),
            guide.get('description', ''),
            tips
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        row_idx += 1

    # 设置列宽
    col_widths = [15, 20, 18, 12, 15, 50, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 设置行高
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 60


def _create_budget_sheet(wb, roadbook_data):
    """预算明细 Sheet"""
    ws = wb.create_sheet("预算明细")

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["类别", "项目", "金额", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    budget = roadbook_data.get('budget', {})
    row_idx = 2

    # 交通费用
    trans = budget.get('transportation', {})
    if trans:
        ws.cell(row=row_idx, column=1, value="交通").border = thin_border
        ws.cell(row=row_idx, column=2, value=f"油费 ({trans.get('fuel_total_liters', 0)}L)").border = thin_border
        ws.cell(row=row_idx, column=3, value=trans.get('fuel_cost', 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value=f"油耗{roadbook_data.get('basic_info', {}).get('fuel_consumption', '')}").border = thin_border
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="交通").border = thin_border
        ws.cell(row=row_idx, column=2, value="高速费").border = thin_border
        ws.cell(row=row_idx, column=3, value=trans.get('toll_fees', 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value="").border = thin_border
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="交通").border = thin_border
        ws.cell(row=row_idx, column=2, value="交通小计").border = thin_border
        cell = ws.cell(row=row_idx, column=3, value=trans.get('transportation_total', 0))
        cell.font = Font(bold=True)
        cell.border = thin_border
        ws.cell(row=row_idx, column=4, value="").border = thin_border
        row_idx += 1
        row_idx += 1  # 空行

    # 住宿费用
    accommodations = budget.get('accommodation', [])
    if accommodations:
        for acc in accommodations:
            total = acc.get('price_per_room', 0) * acc.get('nights', 0) * acc.get('rooms', 0)
            ws.cell(row=row_idx, column=1, value="住宿").border = thin_border
            ws.cell(row=row_idx, column=2, value=f"{acc.get('location', '')} - {acc.get('type', '')}").border = thin_border
            ws.cell(row=row_idx, column=3, value=total).border = thin_border
            ws.cell(row=row_idx, column=4, value=f"{acc.get('nights', 0)}晚 x {acc.get('rooms', 0)}间 x {acc.get('price_per_room', 0)}元").border = thin_border
            row_idx += 1
        row_idx += 1  # 空行

    # 餐饮费用
    food = budget.get('food', {})
    if food:
        ws.cell(row=row_idx, column=1, value="餐饮").border = thin_border
        ws.cell(row=row_idx, column=2, value=f"餐饮费用（{food.get('days', 0)}天）").border = thin_border
        ws.cell(row=row_idx, column=3, value=food.get('group_total', 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value=f"人均{food.get('daily_budget_per_person', 0)}元/天").border = thin_border
        row_idx += 1
        row_idx += 1  # 空行

    # 门票费用
    tickets = budget.get('tickets', [])
    if tickets:
        for ticket in tickets:
            ws.cell(row=row_idx, column=1, value="门票").border = thin_border
            ws.cell(row=row_idx, column=2, value=ticket.get('item', '')).border = thin_border
            ws.cell(row=row_idx, column=3, value=ticket.get('total', 0)).border = thin_border
            ws.cell(row=row_idx, column=4, value=ticket.get('remark', '')).border = thin_border
            row_idx += 1
        row_idx += 1  # 空行

    # 杂费
    misc = budget.get('misc', [])
    if misc:
        for m in misc:
            ws.cell(row=row_idx, column=1, value="杂费").border = thin_border
            ws.cell(row=row_idx, column=2, value=m.get('item', '')).border = thin_border
            ws.cell(row=row_idx, column=3, value=m.get('total', 0)).border = thin_border
            ws.cell(row=row_idx, column=4, value="").border = thin_border
            row_idx += 1
        row_idx += 1  # 空行

    # 总计
    grand_total = budget.get('grand_total', {})
    if grand_total:
        ws.cell(row=row_idx, column=1, value="总计").border = thin_border
        ws.cell(row=row_idx, column=2, value="总费用").border = thin_border
        cell = ws.cell(row=row_idx, column=3, value=grand_total.get('group_total', 0))
        cell.font = Font(bold=True, size=12)
        cell.border = thin_border
        ws.cell(row=row_idx, column=4, value=f"人均 {grand_total.get('per_person', 0)} 元").border = thin_border

    # 设置列宽
    col_widths = [10, 30, 12, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _create_checklist_sheet(wb, roadbook_data):
    """出行清单 Sheet"""
    ws = wb.create_sheet("出行清单")

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    headers = ["分类", "物品清单"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据行
    checklist = roadbook_data.get('checklist', [])
    row_idx = 2
    for category in checklist:
        category_name = category.get('category', '')
        items = category.get('items', [])

        if items:
            start_row = row_idx
            end_row = row_idx + len(items) - 1

            # 写入分类名（先写值再合并）
            ws.cell(row=start_row, column=1, value=category_name)
            ws.cell(row=start_row, column=1).border = thin_border
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical='top')
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

            for i, item in enumerate(items):
                cell = ws.cell(row=row_idx + i, column=2, value=item)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        else:
            ws.cell(row=row_idx, column=1, value=category_name).border = thin_border
            ws.cell(row=row_idx, column=2, value="").border = thin_border

        row_idx += len(items) if items else 1

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50


def get_excel_filename(roadbook_data):
    """根据路书数据生成 Excel 文件名"""
    basic_info = roadbook_data.get('basic_info', {})
    title = basic_info.get('title', '路书')
    start_date = basic_info.get('travel_date_start', '')

    # 清理文件名中的非法字符
    title = title.replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-')

    if start_date:
        return f"路书_{title}_{start_date}.xlsx"
    else:
        return f"路书_{title}.xlsx"
